import time
from sys import platform
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from gologin import GoLogin
import requests
import json
import csv
import openpyxl
import information



InputDoanProxyBatDau=int(input('Đoạn Proxy bắt đầu: '))
InputDoanProxyKetThuc=int(input('Đoạn Proxy kết thúc: '))  
print(str(InputDoanProxyBatDau)+"-->"+str(InputDoanProxyKetThuc))

wb=openpyxl.load_workbook(information.my_InforProfile_path)
sheet_data=wb['Data']

# rangetemp=[1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19]
rangetemp=range(1,35)



for InputDoanProxy in range(InputDoanProxyBatDau,InputDoanProxyKetThuc+1):

	# InputDoanProxy=int(input('Đoạn Proxy: '))

	print("Bạn đang chạy đoạn proxy: "+str(InputDoanProxy))

	doanproxy=information.get_doan_proxy(InputDoanProxy)

	range='A{}:B{}'.format(doanproxy[0],doanproxy[1])
	for row in sheet_data[range]:
		for cell in row:
			cell.value = None

	header=information.headers[InputDoanProxy-1]
	
	data=list()
	for i in rangetemp:
		request=information.API_URL + '/browser/v2?page='+ str(i)
		print(request)
		a=requests.get(request, headers=header)
		dataTemp=json.loads(a.content.decode('utf-8'))
		dataTemp1=dataTemp['profiles']
		data.extend(dataTemp1)

	row=doanproxy[0]
	for x in data:
		proxy=x['proxy']
		name=x['id']

		host=proxy['host']
		sheet_data.cell(row,1).value=host
		sheet_data.cell(row,2).value=name

		row=row+1
	print("Bạn chạy xong đoạn proxy: "+str(InputDoanProxy))
	

wb.close()
wb.save(information.my_InforProfile_path)
